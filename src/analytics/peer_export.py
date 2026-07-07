import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.config import DATABASE_PATH
BASE_DIR = Path(__file__).resolve().parents[2]

OUTPUT_PATH = BASE_DIR / "output"
OUTPUT_PATH.mkdir(exist_ok=True)

EXCEL_FILE = OUTPUT_PATH / "peer_comparison.xlsx"

conn = sqlite3.connect(DATABASE_PATH)

def load_data():
    query = """
    SELECT
        p.company_id,
        c.company_name,
        p.peer_group_name,
        p.year,
        p.metric,
        p.value,
        p.percentile_rank,
        pg.is_benchmark
    FROM peer_percentiles p

    LEFT JOIN companies c
        ON p.company_id = c.id

    LEFT JOIN peer_groups pg
        ON p.company_id = pg.company_id
       AND p.peer_group_name = pg.peer_group_name
    """
    return pd.read_sql(query, conn)

def export_peer_report():
    df = load_data()

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        for group in sorted(df["peer_group_name"].dropna().unique()):
            print(f"Generating Sheet : {group}")
            temp = df[df["peer_group_name"] == group].copy()
            report = temp.pivot_table(
                index=["company_id", "company_name", "year"],
                columns="metric",
                values=["value", "percentile_rank"],
                aggfunc="first"
            )
            report.columns = [f"{a}_{b}" for a, b in report.columns]
            report.reset_index(inplace=True)
            # Add Benchmark Flag
            benchmark = (
                temp.groupby("company_id")["is_benchmark"]
                .max()
                .reset_index()
            )

            report = report.merge(
                benchmark,
                on="company_id",
                how="left"
            )
            sheet_name = group[:31]
            # Add Median Row
            numeric_cols = report.select_dtypes(include="number").columns
            median_row = {}
            for col in report.columns:
                if col in numeric_cols:
                    median_row[col] = report[col].median()
                else:
                    median_row[col] = ""
            median_row["company_name"] = "Peer Group Median"
            report.loc[len(report)] = median_row
            report.to_excel(writer, sheet_name=sheet_name, index=False)
            
            

    # Apply Excel Formatting
    wb = load_workbook(EXCEL_FILE)
    gold = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    red = PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        if "is_benchmark" in headers:
            bench_col = headers.index("is_benchmark") + 1
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row_idx, bench_col).value == 1:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row_idx, col_idx).fill = gold
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 0.75:
                        cell.fill = green
                    elif cell.value >= 0.25:
                        cell.fill = yellow
                    elif cell.value >= 0:
                        cell.fill = red

    wb.save(EXCEL_FILE)
    print("Excel formatting applied successfully!")
    print("\n===================================")
    print("peer_comparison.xlsx Generated")
    print("Location :", EXCEL_FILE)
    print("===================================")
    return None
if __name__ == "__main__":

    export_peer_report()

    conn.close()